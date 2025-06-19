## backend/app/utils/excel.py

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from app.models.estimate import Estimate
from collections import defaultdict


def generate_excel(estimate: Estimate) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = estimate.name[:31].replace(":", "-")

    status_map = {
        "draft": "Черновик",
        "sent": "Отправлена",
        "approved": "Одобрена",
        "paid": "Оплачена",
        "cancelled": "Отменена",
    }
    currency_format = "₽#,##0.00"
    bold_font = Font(bold=True)
    title_font = Font(size=14, bold=True)
    header_fill = PatternFill(
        start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"
    )
    total_fill = PatternFill(
        start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
    )
    cat_total_fill = PatternFill(
        start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    title_cell = ws.cell(row=1, column=1, value=estimate.name)
    title_cell.font = title_font
    title_cell.alignment = center_align

    fields = [
        ("Клиент", estimate.client.name if estimate.client else "—"),
        (
            "Компания клиента",
            (
                estimate.client.company
                if estimate.client and estimate.client.company
                else "—"
            ),
        ),
        (
            "Контакт",
            estimate.client.email if estimate.client and estimate.client.email else "—",
        ),
        ("Статус", status_map.get(estimate.status.value, estimate.status.value)),
        ("Ответственный", estimate.responsible),
        (
            "Дата и время проведения мероприятия",
            estimate.event_datetime if estimate.event_datetime else "—",
        ),
        (
            "Место проведения мероприятия",
            estimate.event_place if estimate.event_place else "—",
        ),
        (
            "НДС",
            f"Включён ({estimate.vat_rate}%)" if estimate.vat_enabled else "Не включён",
        ),
        ("Дата создания", estimate.date.strftime("%d.%m.%Y %H:%M:%S")),
    ]

    row = 3
    for label, value in fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"A{row}"].font = bold_font
        row += 1

    row += 1  # Отступ перед услугами

    # # Красивая область примечаний отдельным блоком
    # if getattr(estimate, "notes", None) and estimate.notes:
    #     ws[f"A{row}"] = "Примечания"
    #     ws[f"A{row}"].font = bold_font
    #     for n in estimate.notes:
    #         row += 1
    #         note_text = (
    #             f"{n.text} — {n.user.name} ({n.created_at.strftime('%d.%m.%Y %H:%M')})"
    #         )
    #         cell = ws[f"B{row}"]
    #         cell.value = note_text
    #         cell.alignment = Alignment(wrap_text=True, vertical="top")
    #     # выделить блок примечаний фоном, если нужно
    #     # for r in range(first_row+1, row+1): ws[f"B{r}"].fill = PatternFill(...)
    #     row += 1

    ws[f"A{row}"] = "Услуги"
    ws[f"A{row}"].font = Font(size=12, bold=True)
    row += 1

    if estimate.use_internal_price:
        headers = [
            "Категория",
            "Название",
            "Описание",
            "Кол-во",
            "Ед. изм.",
            "Внутр. цена",
            "Итог (вн.)",
            "Внеш. цена",
            "Итог (внеш.)",
        ]
    else:
        headers = [
            "Категория",
            "Название",
            "Описание",
            "Кол-во",
            "Ед. изм.",
            "Внеш. цена",
            "Итог (внеш.)",
        ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border
        cell.fill = header_fill

    row += 1
    grouped = defaultdict(list)
    for item in estimate.items:
        grouped[item.category or "Без категории"].append(item)

    # Для финального автоподбора ширины — храним все строки по столбцам:
    all_col_values = [[] for _ in headers]

    internal_service_rows = []
    external_service_rows = []

    for category, items in grouped.items():

        cat_internal_rows = []
        cat_external_rows = []

        for item in items:
            if estimate.use_internal_price:
                values = [
                    category,
                    item.name,
                    item.description,
                    item.quantity,
                    item.unit,
                    item.internal_price,
                    None,
                    item.external_price,
                    None,
                ]
            else:
                values = [
                    category,
                    item.name,
                    item.description,
                    item.quantity,
                    item.unit,
                    item.external_price,
                    None,
                ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = center_align if col not in (2, 3) else left_align
                if estimate.use_internal_price:
                    if col in (6, 8):
                        cell.number_format = currency_format
                    if col == 7:
                        cell.value = f"=F{row}*D{row}"
                        cell.number_format = currency_format
                        cat_internal_rows.append(row)
                        internal_service_rows.append(row)
                    if col == 9:
                        cell.value = f"=H{row}*D{row}"
                        cell.number_format = currency_format
                        cat_external_rows.append(row)
                        external_service_rows.append(row)
                else:
                    if col == 6:
                        cell.number_format = currency_format
                    if col == 7:
                        cell.value = f"=F{row}*D{row}"
                        cell.number_format = currency_format
                        cat_external_rows.append(row)
                        external_service_rows.append(row)

                s = str(val) if val is not None else ""
                if col - 1 < len(all_col_values):
                    all_col_values[col - 1].append(s)
            row += 1

        # Итоги по категории
        ws.cell(row=row, column=1, value=f"Итог по категории: {category}").font = bold_font
        ws.cell(row=row, column=1).fill = cat_total_fill
        if estimate.use_internal_price:
            ws.cell(row=row, column=6, value="Внутр.").font = bold_font
            ws.cell(
                row=row,
                column=7,
                value=f"=SUM(G{cat_internal_rows[0]}:G{cat_internal_rows[-1]})",
            ).font = bold_font
            ws.cell(row=row, column=7).number_format = currency_format
            ws.cell(row=row, column=7).fill = cat_total_fill
            ws.cell(row=row, column=8, value="Внешн.").font = bold_font
            ws.cell(
                row=row,
                column=9,
                value=f"=SUM(I{cat_external_rows[0]}:I{cat_external_rows[-1]})",
            ).font = bold_font
            ws.cell(row=row, column=9).number_format = currency_format
            ws.cell(row=row, column=9).fill = cat_total_fill
            row += 2
        else:
            ws.cell(row=row, column=6, value="Итог").font = bold_font
            ws.cell(
                row=row,
                column=7,
                value=f"=SUM(G{cat_external_rows[0]}:G{cat_external_rows[-1]})",
            ).font = bold_font
            ws.cell(row=row, column=7).number_format = currency_format
            ws.cell(row=row, column=7).fill = cat_total_fill
            row += 2
        # Отступ после категории
        row += 1

    # Финальные итоги
    row += 1
    ws[f"E{row}"] = "Итого по всем категориям:"
    ws[f"E{row}"].font = bold_font
    ws[f"E{row}"].fill = total_fill
    ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    if estimate.use_internal_price:
        ws[f"F{row}"] = "Внутр."
        ws[f"G{row}"] = (
            f"=SUM({','.join([f'G{r}' for r in internal_service_rows])})"
            if internal_service_rows
            else 0
        )
        ws[f"G{row}"].number_format = currency_format
        ws[f"G{row}"].font = bold_font
        ws[f"G{row}"].fill = total_fill

        ws[f"H{row}"] = "Внешн."
        ws[f"I{row}"] = (
            f"=SUM({','.join([f'I{r}' for r in external_service_rows])})"
            if external_service_rows
            else 0
        )
        ws[f"I{row}"].number_format = currency_format
        ws[f"I{row}"].font = bold_font
        ws[f"I{row}"].fill = total_fill

        # Маржа
        row += 1
        ws[f"E{row}"] = "Маржа:"
        ws[f"E{row}"].font = bold_font
        ws[f"E{row}"].fill = total_fill
        ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"I{row}"] = f"=I{row-1}-G{row-1}"
        ws[f"I{row}"].number_format = currency_format
        ws[f"I{row}"].font = bold_font
        ws[f"I{row}"].fill = total_fill
    else:
        ws[f"F{row}"] = "Внешн."
        ws[f"G{row}"] = (
            f"=SUM({','.join([f'G{r}' for r in external_service_rows])})"
            if external_service_rows
            else 0
        )
        ws[f"G{row}"].number_format = currency_format
        ws[f"G{row}"].font = bold_font
        ws[f"G{row}"].fill = total_fill

    # НДС
    if estimate.vat_enabled:
        row += 1
        ws[f"E{row}"] = f"НДС ({estimate.vat_rate}%)"
        ws[f"E{row}"].font = bold_font
        ws[f"E{row}"].fill = total_fill
        ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"I{row}"] = f"=I{row-1}*{estimate.vat_rate/100}"
        ws[f"I{row}"].number_format = currency_format
        ws[f"I{row}"].font = bold_font
        ws[f"I{row}"].fill = total_fill

    # Итог с НДС
    row += 1
    ws[f"E{row}"] = "Итого с НДС"
    ws[f"E{row}"].font = bold_font
    ws[f"E{row}"].fill = total_fill
    ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    if estimate.vat_enabled:
        ws[f"I{row}"] = f"=I{row-1}+I{row-2}"
    else:
        ws[f"I{row}"] = f"=I{row-1}"
    ws[f"I{row}"].number_format = currency_format
    ws[f"I{row}"].font = bold_font
    ws[f"I{row}"].fill = total_fill

    # Автоподбор ширины по max длине текста по всему документу
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # если многострочный текст, разбиваем
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                # можно чуть увеличить запас для чисел и bold текста
                if cell.font and cell.font.bold:
                    longest = int(longest * 1.15)
                max_length = max(max_length, longest)
        ws.column_dimensions[col_letter].width = max(min(max_length + 3, 33), 10)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
