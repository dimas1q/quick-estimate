from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from app.schemas.analytics import GlobalAnalytics


def generate_analytics_excel(ga: GlobalAnalytics) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics"

    bold = Font(bold=True)

    ws.append(["Метрика", "Значение"])
    ws["A1"].font = ws["B1"].font = bold
    ws.append(["Всего смет", ga.total_estimates])
    ws.append(["Общая сумма", ga.total_amount])
    ws.append(["Средняя сумма", ga.average_amount])
    ws.append(["Медиана по сметам", ga.median_amount])
    ws.append(["ARPU", ga.arpu])
    ws.append(["MoM рост (%)", ga.mom_growth or 0])
    ws.append(["YoY рост (%)", ga.yoy_growth or 0])

    ws.append([])
    ws.append(["Период", "Сумма"])
    start = ws.max_row
    ws[f"A{start}"].font = ws[f"B{start}"].font = bold
    for item in ga.timeseries:
        ws.append([item.period, item.value])

    ws.append([])
    ws.append(["Top-10 клиентов", "Выручка"])
    start = ws.max_row
    ws[f"A{start}"].font = ws[f"B{start}"].font = bold
    for cli in ga.top_clients:
        ws.append([cli.name, cli.total_amount])

    ws.append([])
    ws.append(["Ответственный", "Число смет", "Выручка"])
    start = ws.max_row
    ws[f"A{start}"].font = ws[f"B{start}"].font = ws[f"C{start}"].font = bold
    for r in ga.by_responsible:
        ws.append([r.name, r.estimates_count, r.total_amount])

    ws.append([])
    ws.append(["Top-10 услуг", "Выручка"])
    start = ws.max_row
    ws[f"A{start}"].font = ws[f"B{start}"].font = bold
    for srv in ga.top_services:
        ws.append([srv.name, srv.total_amount])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
